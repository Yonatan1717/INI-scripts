import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
import ipaddress

DEFAULT_IPSEC_PSK = "DMVPN-KEY"
SSH_DOMAIN = "lab.local"


def read_sheet(filename, sheet):
    md_columns = {"site", "intf_prefix", "router-id", "brukernavn", "passord", "vty_lines",}
    tunnel_columns = {
        "tunnel id",
        "gre mode",
        "ip address",
        "mask",
        "vrf",
        "source",
        "network-id",
        "destination",
        "ipsec",
        "ipsec key",
    }

    return {
        "md": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols=lambda c: str(c).strip().lower() in md_columns,
            skiprows=0,
            nrows=2
        ),
        "ip_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:J",
            skiprows=3,
            nrows=4
        ),
        "vrf_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:D",
            skiprows=8,
            nrows=4
        ),
        "tunnel_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols=lambda c: str(c).strip().lower() in tunnel_columns,
            skiprows=13,
            nrows=2
        )
    }


def is_true(value):
    """
    Godta True/TRUE/1/yes/ja/x fra sheets.
    """
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value == 1

    return str(value).strip().lower() in {"true", "1", "yes", "ja", "y", "x"}


def getNetId(ip, mask):
    ip = ip.split(".")
    mask = mask.split(".")
    netid = []
    wild_mask = [255, 255, 255, 255]
    wild = []

    for i in range(4):
        ip_b = int(ip[i])
        ip_m = int(mask[i])
        ip_w = wild_mask[i]

        wild.append(str(ip_m ^ ip_w))
        netid.append(str(ip_b & ip_m))

    return ".".join(netid), ".".join(mask), ".".join(wild)


def create_vrf(vrf_data, sn):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    for index, row in vrf_data.iterrows():
        vrf_name = row["vrf"]
        vrf_rt = row["rt"]
        vrf_rd = row["rt"].replace(":", f":{sn}")
        vrf_loopback = row["loopback"]
        vrf_laddr = row["laddr"]

        my_data["network_info"][vrf_name] = {
            "rt": vrf_rt,
            "loopback": vrf_loopback,
            "laddr": vrf_laddr
        }

        vrf_s = []
        vrf_s.append(f"rd {vrf_rd}")
        vrf_s.append(f"route-target export {vrf_rt}")
        vrf_s.append(f"route-target import {vrf_rt}")
        vrf_s.append("exit")
        my_data["config"][f"ip vrf {vrf_name}"] = vrf_s

        intf_s = []
        intf_s.append(f"ip vrf forwarding {vrf_name}")
        intf_s.append(f"ip address {vrf_laddr} 255.255.255.255")
        intf_s.append("exit")
        my_data["config"][f"interface loopback{vrf_loopback}"] = intf_s

    return my_data


def create_interface(ip_data, intf_prefix):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}
    intf_nums = list(ip_data["interface"])
    
    tot_pri = ip_data["pri-1-10"].sum()
    pri_prc_for_class_default = 25
    pri_left = 100 - pri_prc_for_class_default
    pol_maps= {}

    for index, row in ip_data.iterrows():
        vrf = row["vrf"]
        vlan = row["vlan"]
        pri = row["pri-1-10"]
        pri_prc = int((pri / tot_pri) * pri_left)

        intf = row["interface"]
        sub = True if intf_nums.count(intf) > 1 else False

        ip_address = row["address min"]
        mask = row["mask"]

        my_data["network_info"][
            f"{intf_prefix}{intf}.{vlan}" if sub else f"{intf_prefix}{intf}"
        ] = {
            "vrf": vrf,
            "vlan": vlan,
            "interface": intf,
            "sub": sub,
            "address": ip_address,
            "mask": mask,
            "pri_prc": pri_prc,
        }

        intf_s = []

        if sub:
            intf_s.append(f"encapsulation dot1Q {vlan}")

        intf_s.append(f"ip vrf forwarding {vrf}")
        intf_s.append(f"ip address {ip_address} {mask}")
        intf_s.append("no shutdown")
        intf_s.append("exit")

        my_data["config"][
            f"interface {intf_prefix}{intf}.{vlan}" if sub
            else f"interface {intf_prefix}{intf}"
        ] = intf_s

        if sub:
            my_data["config"][f"class-map match-any QRS-CLASS-{vlan}"] = [
                f"match input-interface {intf_prefix}{intf}.{vlan}",
                "exit"
            ]

            if f"policy-map QRS-SITE-POLICY" not in pol_maps:
                pol_maps[f"policy-map QRS-SITE-POLICY"] = []

            pol_maps[f"policy-map QRS-SITE-POLICY"].append(
                {
                    f"class QRS-CLASS-{vlan}": [
                        f"bandwidth percent {pri_prc}", 
                        "exit"
                    ]
                }   
            )

    intf_s = []
    intf_s.append("no shutdown")
    intf_s.append("exit")
    my_data["config"][f"interface {intf_prefix}{intf}"] = intf_s
    
    if f"policy-map QRS-SITE-POLICY" in pol_maps:
        my_data["config"].update(pol_maps)
        
        my_data["config"][f"\ninterface {intf_prefix}1"] = [
            "service-policy output QRS-SITE-POLICY",
            "exit"
        ]

        

    return my_data


def create_mp_bgp_config(vrf_data, tunnel_data, ip_data, sites_data, router_id, site_number):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    bgp_s = []
    vpnv4_s = []

    rt = list(vrf_data.iterrows())[0][1]["rt"]
    as_num = rt.split(":")[0]

    tmp = f"neighbor {router_id} remote-as {as_num}"
    tmp2 = f"neighbor {router_id} update-source loopback0"
    vpn_tmp = f"neighbor {router_id} activate"
    vpn_tmp2 = f"neighbor {router_id} send-community extended"

    other_sites = sites_data.copy()

    if f"site {site_number}" in other_sites:
        del other_sites[f"site {site_number}"]

    if "hub" in other_sites:
        del other_sites["hub"]

    for site, site_data in other_sites.items():
        net_info = site_data["network_info"]
        loop0 = net_info["loopback0"]

        bgp_s.append(f"neighbor {loop0['address']} remote-as {as_num}")
        bgp_s.append(f"neighbor {loop0['address']} update-source loopback0")

        vpnv4_s.append(f"neighbor {loop0['address']} activate")
        vpnv4_s.append(f"neighbor {loop0['address']} send-community extended")

        # Update BGP-neighbor-konfigurasjonen på allerede genererte sites.
        l = sites_data[site]["config"][f"router bgp {as_num}"][:-5]
        sites_data[site]["config"][f"router bgp {as_num}"] = (
            sites_data[site]["config"][f"router bgp {as_num}"][-5:]
        )

        l.insert(0, tmp)
        l.insert(1, tmp2)
        l = list(set(l))
        l.sort()

        for i, x in enumerate(l):
            sites_data[site]["config"][f"router bgp {as_num}"].insert(i, x)

        l = sites_data[site]["config"][f"router bgp {as_num}"][-5][
            "address-family vpnv4"
        ][:-1]

        l.insert(0, vpn_tmp)
        l.insert(0, vpn_tmp2)
        l = list(set(l))
        l.sort()
        l.append("exit-address-family")

        sites_data[site]["config"][f"router bgp {as_num}"][-5][
            "address-family vpnv4"
        ] = l

    vpnv4_s.append("exit-address-family")
    bgp_s.append({"address-family vpnv4": vpnv4_s})

    for index, row in ip_data.iterrows():
        ipv4_s = []
        vrf = row["vrf"]

        is_tunnel = tunnel_data[tunnel_data["vrf"] == vrf].shape[0] > 0
        
        
        vrf_loop_addr = vrf_data[vrf_data["vrf"] == vrf]["laddr"].values[0]
        vrf_loop_mask = "255.255.255.255"

        network = row["nett id"]
        mask = row["mask"]

        if not is_tunnel:
            ipv4_s.append(f"network {network} mask {mask}")

     
        ipv4_s.append(f"network {vrf_loop_addr} mask {vrf_loop_mask}")
        ipv4_s.append("exit-address-family")

        bgp_s.append({f"address-family ipv4 vrf {vrf}": ipv4_s})

    bgp_s.append("exit")
    my_data["config"][f"router bgp {as_num}"] = bgp_s

    return my_data, sites_data


def create_ipsec_config(network_id, vrf, psk=DEFAULT_IPSEC_PSK):
    """
    Lager IPsec-konfigurasjon for en DMVPN-tunnel.
    """
    suffix = str(network_id).strip()

    proposal = f"DMVPN-IKEV2-PROP-{suffix}"
    policy = f"DMVPN-IKEV2-POL-{suffix}"
    keyring = f"DMVPN-IKEV2-KR-{suffix}"
    ikev2_profile = f"DMVPN-IKEV2-PROFILE-{suffix}"
    transform_set = f"DMVPN-TS-{suffix}"
    ipsec_profile = f"DMVPN-IPSEC-{suffix}"

    config = {}

    config[f"crypto ikev2 proposal {proposal}"] = [
        "encryption aes-cbc-256",
        "integrity sha256",
        "group 14",
        "exit",
    ]

    config[f"crypto ikev2 policy {policy}"] = [
        f"match fvrf {vrf}",
        f"proposal {proposal}",
        "exit",
    ]

    config[f"crypto ikev2 keyring {keyring}"] = [
        {
            "peer ANY": [
                "address 0.0.0.0 0.0.0.0",
                f"pre-shared-key local {psk}",
                f"pre-shared-key remote {psk}",
                "exit",
            ]
        },
        "exit",
    ]

    config[f"crypto ikev2 profile {ikev2_profile}"] = [
        f"match fvrf {vrf}",
        "match identity remote address 0.0.0.0 0.0.0.0",
        "authentication remote pre-share",
        "authentication local pre-share",
        f"keyring local {keyring}",
        "exit",
    ]

    config[
        f"crypto ipsec transform-set {transform_set} esp-aes 256 esp-sha256-hmac"
    ] = [
        "mode transport",
        "exit",
    ]

    config[f"crypto ipsec profile {ipsec_profile}"] = [
        f"set transform-set {transform_set}",
        f"set ikev2-profile {ikev2_profile}",
        "exit",
    ]

    return config, ipsec_profile


def create_tunnel_config(tunnel_data, sites_data: dict, is_hub: bool):
    my_data = {}
    my_data["network_info"] = {}
    my_data["config"] = {}

    if is_hub:
        hub_data = {}
    else:
        hub = sites_data["hub"]
        hub_data = sites_data[hub]["network_info"]

    for idx, row in tunnel_data.iterrows():
        tunnel = row["tunnel id"]
        mode = row["gre mode"]
        ip_address = row["ip address"]
        mask = row["mask"]
        vrf = row["vrf"]
        source = row["source"]
        network_id = row["network-id"]

        ipsec_enabled = is_true(row.get("ipsec", False))

        # Valgfri egen PSK per tunnel.
        # om kollone tom brukes DEFAULT_IPSEC_PSK.
        psk = row.get("ipsec key", DEFAULT_IPSEC_PSK)
        if pd.isna(psk) or str(psk).strip() == "":
            psk = DEFAULT_IPSEC_PSK
        else:
            psk = str(psk).strip()

        if mode != "multipoint" and "destination" in row.index:
            destination = row["destination"]

        tunnel_info = {
            "is hub": is_hub,
            "vrf": vrf,
            "ip address": ip_address,
            "mask": mask,
            "source": source,
            "tunnel id": tunnel,
            "mode": mode,
            "ipsec": ipsec_enabled,
        }

        my_data["network_info"][tunnel] = tunnel_info

        tun_s = []
        tun_s.append(f"ip vrf forwarding {vrf}")
        tun_s.append(f"qos pre-classify")
        tun_s.append(f"ip address {ip_address} {mask}")
        tun_s.append(f"tunnel source {source}")
        tun_s.append(f"tunnel vrf {vrf}")

        if mode == "multipoint":
            tun_s.append(f"tunnel mode gre {mode}")

            if not is_hub:
                hub_tun_ip = hub_data.get(tunnel, {}).get("ip address", "")
                hub_source = hub_data.get(tunnel, {}).get("source", "")

                tun_s.append(f"ip nhrp map {hub_tun_ip} {hub_source}")
                tun_s.append(f"ip nhrp map multicast {hub_source}")
                tun_s.append(f"ip nhrp nhs {hub_tun_ip}")
            else:
                tun_s.append("ip nhrp map multicast dynamic")

            tun_s.append(f"ip nhrp network-id {network_id}")

        else:
            print(f"Mode må være multipoint for tunnel {tunnel}")

        # IPsec aktiveres bare når kolonnen 'ipsec' er TRUE/1/yes/ja/x.
        if ipsec_enabled:
            ipsec_config, ipsec_profile = create_ipsec_config(network_id, vrf, psk)
            my_data["config"].update(ipsec_config)
            tun_s.append(f"tunnel protection ipsec profile {ipsec_profile}")

        tun_s.append("exit")

        my_data["config"][f"interface {tunnel}"] = tun_s
        my_data["network_info"][tunnel] = tunnel_info

    return my_data


def create_tunnel_eigrp_config(vrf_data, tunnel_data, ip_data, is_hub):
    my_data = {}
    my_data["network_info"] = {}
    my_data["config"] = {}

    vrfs = {}
    networks = []

    for idx, row in tunnel_data.iterrows():
        network_id = row["network-id"]
        tun = row["tunnel id"]

        vrf = row["vrf"]
        vrfs[vrf] = []

        ip = row["ip address"]
        mask = row["mask"]

        netinfo = getNetId(ip, mask)
        networks.append(netinfo)

        vrfs[vrf].insert(0, network_id)
        vrfs[vrf].insert(1, tun)
        vrfs[vrf].append(netinfo)

    for idx, row in ip_data.iterrows():
        if row["vrf"] in vrfs:
            network = row["nett id"]
            mask = row["mask"]
            vrfs[row["vrf"]].append(getNetId(network, mask))

    tun_s = {}

    for vrf, nets in vrfs.items():
        tun_vrf_s = []

        for net in nets[2:]:
            if isinstance(net, tuple):
                network, mask, wild = net
                tun_vrf_s.append(f"network {network} {wild}")
            else:
                tun_vrf_s.append(f"network {net}")

        if is_hub:
            tun_vrf_s.append(f"af-interface {nets[1]}")
            tun_vrf_s.append("no split-horizon")
            tun_vrf_s.append("no next-hop-self")
            tun_vrf_s.append("exit-af-interface ")

        tun_vrf_s.append("exit-address-family")
        tun_s[
            f"address-family ipv4 vrf {vrf} autonomous-system {nets[0]}"
        ] = tun_vrf_s

    tun_s["exit"] = []
    my_data["config"]["router eigrp DMVPN-EIGRP"] = tun_s

    return my_data


def enable_ssh(md, domain=SSH_DOMAIN):
    """
    Genererer lokal SSH-konfig
    brukernavn + passord.
    """
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    if md.empty:
        return my_data

    row = md.iloc[0]

    username = row.get("brukernavn", "")
    password = row.get("passord", "")
    vty_lines = row.get("vty_lines", 6)

    if pd.isna(username) or pd.isna(password):
        return my_data

    username = str(username).strip()
    password = str(password).strip()

    if not username or not password:
        return my_data

    my_data["config"][f"ip domain name {domain}"] = []
    my_data["config"][
        f"username {username} privilege 15 secret {password}"
    ] = []
    my_data["config"]["crypto key generate rsa general-keys modulus 4096"] = []
    my_data["config"]["ip ssh version 2"] = []
    my_data["config"][f"line vty {' '.join(x.strip(' ') for x in vty_lines.split('-'))}"] = [
        "login local",
        "transport input ssh",
        "exit",
    ]

    return my_data


def fetch_site_data(config_file, site_number):
    try:
        data = {}
        hub = ""
        is_hub = False

        with open(config_file, "r") as f:
            try:
                data = json.load(f)
                hub = data.get("hub", "")
            except json.JSONDecodeError:
                data = {}
                data["hub"] = f"site {site_number}"
                hub = f"site {site_number}"

    except FileNotFoundError:
        data = {}
        data["hub"] = f"site {site_number}"
        hub = f"site {site_number}"

    return data, hub == f"site {site_number}"


def create_global_config(router_id, intf_prefix, sn):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    my_data["config"][f"hostname RS{sn}"] = []

    ospf_s = []
    ospf_s.append(f"router-id {router_id}")
    ospf_s.append("exit")
    my_data["config"]["router ospf 1"] = ospf_s

    my_data["config"]["interface loopback0"] = []
    my_data["config"]["interface loopback0"].append(
        f"ip address {router_id} 255.255.255.255"
    )
    my_data["config"]["interface loopback0"].append("ip ospf 1 area 0")
    my_data["config"]["interface loopback0"].append("exit")

    my_data["config"]["mpls ldp router-id Loopback0 force"] = []

    intf_s = []
    intf_s.append("ip address dhcp")
    intf_s.append("ip ospf 1 area 0")
    intf_s.append("mpls ip")
    intf_s.append("no shutdown")
    intf_s.append("exit")
    my_data["config"][f"interface {intf_prefix}1"] = intf_s

    my_data["network_info"]["loopback0"] = {
        "address": router_id,
        "mask": "255.255.255.255"
    }

    return my_data


def set_up_DHCP_for_vrf_lans(ip_data):
    my_config = {}
    my_config["config"] = {}
    my_config["network_info"] = {}

    for idx, row in ip_data.iterrows():
        vrf = row["vrf"]
        ip_gw = row["address min"]
        network = row["nett id"]
        mask = row["mask"]
        num_res = row["antall-res"]
        
        ip_res_to = str(ipaddress.ip_address(ip_gw) + num_res)
        
        my_config["config"][f"ip dhcp pool DHCP-{vrf}"] = [
            f"vrf {vrf}",
            f"network {network} {mask}",
            f"default-router {ip_gw}",
            "exit"
        ]

        my_config["config"][f"ip dhcp excluded-address vrf {vrf} {ip_gw} {ip_res_to}"] = []

        my_config["network_info"][f"DHCP-{vrf}"] = {
            "network": network,
            "mask": mask,
            "default-router": ip_gw,
            "ip_res_to": ip_res_to
        }
        
    return my_config

    
def configure_site(sheet_file, config_file, sheet):
    sheet_data = read_sheet(sheet_file, sheet)

    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    #HENT NØDVENDIG DATA 
    ip_data = sheet_data["ip_data"]
    vrf_data = sheet_data["vrf_data"]
    tunnel_data = sheet_data["tunnel_data"]
    md = sheet_data["md"]

    router_id = md.iloc[0]["router-id"]
    sn = md.iloc[0]["site"]
    intf_prefix = md.iloc[0]["intf_prefix"]

    data, is_hub = fetch_site_data(config_file, sn)
    
    #OPPRETT GLOBAL KONFIGURASJON
    my_data = create_global_config(router_id, intf_prefix, sn)
    
    #SSH
    d_ssh = enable_ssh(md)
    my_data["config"].update(d_ssh["config"])
    
    #VRF
    d_vrf = create_vrf(vrf_data, sn)
    my_data["config"].update(d_vrf["config"])
    my_data["network_info"].update(d_vrf["network_info"])

    #INTERFACE
    d_ip = create_interface(ip_data, intf_prefix)
    my_data["config"].update(d_ip["config"])
    my_data["network_info"].update(d_ip["network_info"])
    
    #DHCP
    d_dhcp = set_up_DHCP_for_vrf_lans(ip_data)
    my_data["config"].update(d_dhcp["config"])
    my_data["network_info"].update(d_dhcp["network_info"])
    
    #BGP
    d_bgp, data = create_mp_bgp_config(vrf_data, tunnel_data, ip_data, data, router_id, sn)
    my_data["config"].update(d_bgp["config"])

    #TUNNEL
    d_tunnel = create_tunnel_config(tunnel_data, data, is_hub)
    my_data["config"].update(d_tunnel["config"])
    my_data["network_info"].update(d_tunnel["network_info"])

    #EIGRP
    d_eigrp = create_tunnel_eigrp_config(vrf_data, tunnel_data, ip_data, is_hub)
    my_data["config"].update(d_eigrp["config"])
    my_data["network_info"].update(d_eigrp["network_info"])
    
    #INTerFACE PREFIX
    my_data["intf_prefix"] = intf_prefix

    #LAGRE
    data[f"site {sn}"] = my_data

    with open(config_file, "w") as f:
        json.dump(data, f, indent=4)
        print(f"Config for site {sn} har blitt lagret i {config_file}")

    return data


def config_to_text(data, indent=0):
    lines = []
    prefix = "    " * indent

    if isinstance(data, dict):
        for key, value in data.items():
            lines.append("!")
            lines.append(prefix + key)
            lines.extend(config_to_text(value, indent + 1))

    elif isinstance(data, list):
        for value in data:
            if isinstance(value, str):
                lines.append(prefix + value)
            else:
                lines.extend(config_to_text(value, indent))

    elif isinstance(data, str):
        lines.append(prefix + data)
        
    lines.append("!")
  
    return lines


def create_or_update_config_files(data):
    del data["hub"]

    for site, site_data in data.items():
        config = site_data["config"]
        text = config_to_text(config)

        with open(
            f"site_text_configs/{site}.txt",
            "w",
            encoding="utf-8"
        ) as f:
            f.write("\n".join(text))
            
        print(f"Text versjon av config for site {site} har blitt lagret i site_text_configs/{site}.txt")


def main():
    if not os.path.exists("site_text_configs"):
        os.makedirs("site_text_configs")

    sheet_file = sys.argv[1]
    config_file = (
        sys.argv[2]
        if len(sys.argv) > 2
        else "EDGE_ROUTER_configs.json"
    )

    sheets = load_workbook(sheet_file).sheetnames

    data = {}

    for sheet in sheets:
        data.update(configure_site(sheet_file, config_file, sheet))
    
    print()

    create_or_update_config_files(data)
    print("\nAlle config har blitt lagret.")


if __name__ == "__main__":
    main()
