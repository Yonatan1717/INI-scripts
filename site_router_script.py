import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook

def read_sheet(filename, sheet):
    return {
        "md": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:B",
            skiprows=0,
            nrows=2
        ),
        "ip_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:H",
            skiprows=3,
            nrows=4
        ),

        "vrf_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:D",
            skiprows=8,
            nrows=4
        ),

        "tunnel_data": pd.read_excel(
            filename,
            sheet_name=sheet,
            usecols="A:H",
            skiprows=13,
            nrows=2
        )
    }

def getNetId(ip, mask):
    ip = ip.split(".")
    mask = mask.split(".")
    netid = []
    wild_mask = [255, 255, 255, 255]
    wild = []
    for i in range(4):
        ip_b = int(ip[i])
        ip_m = int(mask[i])
        ip_w = wild_mask[i]

        wild.append(str(ip_m ^ ip_w))

        netid.append(str(ip_b & ip_m))

    return ".".join(netid), ".".join(mask), ".".join(wild)

def create_vrf(vrf_data, sn):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}
    
    for index, row in vrf_data.iterrows():
        vrf_name = row['vrf']
        vrf_rt = row['rt']
        vrf_rd = row['rt'].replace(":", f":{sn}")
        vrf_loopback = row['loopback']
        vrf_laddr= row['laddr']
        my_data["network_info"][vrf_name] = {
            "rt": vrf_rt,
            "loopback": vrf_loopback,
            "laddr": vrf_laddr
        }

        print(f"lager VRF: {vrf_name} med RD: {vrf_rd}")
        vrf_s = []
        vrf_s.append(f"rd {vrf_rd}")
        vrf_s.append(f"route-target export {vrf_rt}")
        vrf_s.append(f"route-target import {vrf_rt}")
        vrf_s.append(f"exit")
        my_data["config"][f"ip vrf {vrf_name}"] = vrf_s
        intf_s = []
        intf_s.append(f"ip vrf forwarding {vrf_name}")
        intf_s.append(f"ip address {vrf_laddr} 255.255.255.255")
        intf_s.append(f"exit")
        my_data["config"][f"interface loopback{vrf_loopback}"] = intf_s
        
    return my_data

def create_interface(ip_data, intf_prefix):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}
    intf_nums = list(ip_data['interface'])
    
    for index, row in ip_data.iterrows():
        vrf = row['vrf']
        vlan = row['vlan']

        intf = row['interface']
        sub = True if intf_nums.count(intf) > 1 else False

        ip_address = row['address min']
        mask = row['mask']
        my_data["network_info"][f"{intf_prefix}{intf}.{vlan}" if sub else f"{intf_prefix}{intf}"] = {
            "vrf": vrf,
            "vlan": vlan,
            "interface": intf,
            "sub": sub,
            "address": ip_address,
            "mask": mask
        }

        print(f"lager Interface for VRF: {vrf} med IP: {ip_address} og Mask: {mask}")
        intf_s = []
        intf_s.append(f"encapsulation dot1Q {vlan}")
        intf_s.append(f"ip vrf forwarding {vrf}")
        intf_s.append(f"ip address {ip_address} {mask}")
        intf_s.append("no shutdown")
        intf_s.append("exit")
        my_data["config"][f"interface {intf_prefix}{intf}.{vlan}" if sub else f"interface {intf_prefix}{intf}"] = intf_s

        #slå på parre (enable the interface)
        intf_s = []
        intf_s.append("no shutdown")
        intf_s.append("exit")
        my_data["config"][f"interface {intf_prefix}{intf}"] = intf_s

    return my_data

def create_mp_bgp_config(vrf_data, ip_data, sites_data, router_id, site_number):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    bgp_s = []
    vpnv4_s= []
   
    rt = list(vrf_data.iterrows())[0][1]['rt']
         
    as_num = rt.split(":")[0]
        
    tmp = f"neighbor {router_id} remote-as {as_num}"
    tmp2 = f"neighbor {router_id} update-source loopback0"
    vpn_tmp = f"neighbor {router_id} activate"
    vpn_tmp2 = f"neighbor {router_id} send-community extended"
    other_sites = sites_data.copy()
    if f"site {site_number}" in other_sites:
        del other_sites[f"site {site_number}"]
    if "hub" in other_sites:
        del other_sites["hub"]

    for site, site_data in other_sites.items():
        net_info = site_data["network_info"]
        loop0 = net_info[f"loopback0"]
        
        bgp_s.append(f"neighbor {loop0['address']} remote-as {as_num}")
        bgp_s.append(f"neighbor {loop0['address']} update-source loopback0")

        vpnv4_s.append(f"neighbor {loop0['address']} activate")
        vpnv4_s.append(f"neighbor {loop0['address']} send-community extended")

        # update BGP neighbor configuration for each site

        l = sites_data[site]["config"][f"router bgp {as_num}"][:-3]
        sites_data[site]["config"][f"router bgp {as_num}"] = sites_data[site]["config"][f"router bgp {as_num}"][-3:]
        l.insert(0, tmp)
        l.insert(1, tmp2)
        l = list(set(l)) 
        for i,x in enumerate(l):
            sites_data[site]["config"][f"router bgp {as_num}"].insert(0, x)

         
        l = sites_data[site]["config"][f"router bgp {as_num}"][-3][f"address-family vpnv4"][:-1]

        l.insert(0, vpn_tmp)
        l.insert(0, vpn_tmp2)
        l = list(set(l))
        l.append("exit-address-family")

        sites_data[site]["config"][f"router bgp {as_num}"][-3][f"address-family vpnv4"] = l

    vpnv4_s.append("exit-address-family")
    bgp_s.append({f"address-family vpnv4": vpnv4_s})

    for index, row in ip_data.iterrows():
        ipv4_s = []
        vrf = row['vrf']
        if vrf == "UNET":
            continue

        network = row['nett id']
        mask = row['mask']
        ipv4_s.append(f"network {network} mask {mask}")
        ipv4_s.append("exit-address-family")
        
        bgp_s.append({f"address-family ipv4 vrf {vrf}": ipv4_s})


    my_data["config"][f"router bgp {as_num}"] = bgp_s
    
    return my_data, sites_data

def create_tunnel_config(tunnel_data, sites_data: dict, is_hub: bool):
    my_data = {}
    my_data["network_info"] = {}
    my_data["config"] = {}

    if is_hub:
        hub_data = {}
    else:
        hub = sites_data["hub"]
        hub_data = sites_data[hub]["network_info"]

    for idx, row in tunnel_data.iterrows():
        tunnel= row['tunnel id']
        mode = row['gre mode']
        ip_address = row['ip address']
        mask = row['mask']
        vrf = row['vrf']
        source = row['source']
        network_id = row['network-id']

        if mode != "multipoint":
            destination = row['destination']


        tunnel_info = {
            "is hub": is_hub,
            "vrf": vrf,
            "ip address": ip_address,
            "mask": mask,
            "source": source,
            "tunnel id": tunnel,
            "mode": mode,
        }

        my_data["network_info"][tunnel] = tunnel_info

        tun_s = []
        tun_s.append(f"ip vrf forwarding {vrf}")
        tun_s.append(f"ip address {ip_address} {mask}")
        tun_s.append(f"tunnel source Loopback0")

        if mode == "multipoint":
            tun_s.append(f"tunnel mode gre {mode}")
            if not is_hub:
                hub_tun_ip = hub_data.get(tunnel, {}).get("ip address", "")
                hub_source = hub_data.get(tunnel, {}).get("source", "")

                tun_s.append(f"ip nhrp map {hub_tun_ip} {hub_source}")
                tun_s.append(f"ip nhrp map multicast {hub_source}")
                tun_s.append(f"ip nhrp nhs {hub_tun_ip}")
            else:
                tun_s.append(f"ip nhrp map multicast dynamic")


            tun_s.append(f"ip nhrp network-id {network_id}")

        else:
            print(f"Mode må være multipoint for tunnel {tunnel}")

        tun_s.append("exit")
        my_data["config"][f"interface {tunnel}"] = tun_s
        my_data["network_info"][tunnel] = tunnel_info

    return my_data

def create_tunnel_eigrp_config(vrf_data, tunnel_data, ip_data, is_hub):
    my_data = {}
    my_data["network_info"] = {}
    my_data["config"] = {}

    vrfs = {} 
    networks = []
    for idx, row in tunnel_data.iterrows():
        network_id = row["network-id"]
        tun = row["tunnel id"]

        vrf = row["vrf"]
        vrfs[vrf] = [] 
        ip = row["ip address"]
        mask = row["mask"]

        #hent riktig vrf rad fra vrf_data
        vrf_info = vrf_data[vrf_data["vrf"] == vrf]
        laddr = vrf_info["laddr"].values[0]
        lop_mask = "255.255.255.255"


        netinfo = getNetId(ip, mask)
        vrf_loopback_netinfo = getNetId(laddr, lop_mask)

        networks.append(netinfo)
        networks.append(vrf_loopback_netinfo)

        vrfs[vrf].insert(0, network_id)
        vrfs[vrf].insert(1, tun)
        vrfs[vrf].append(netinfo)
        vrfs[vrf].append(vrf_loopback_netinfo)


    for idx, row in ip_data.iterrows():
        if row["vrf"] in vrfs:
            network = row["nett id"]
            mask = row["mask"]
            vrfs[row["vrf"]].append(getNetId(network, mask))

    tun_s = {}
    for vrf, nets in vrfs.items():
        tun_vrf_s = []
        for net in nets[2:]:
            if isinstance(net, tuple):
                network, mask, wild = net
                tun_vrf_s.append(f"network {network} {wild}")
            else:
                tun_vrf_s.append(f"network {net}")

        if is_hub:
            tun_vrf_s.append(f"af-interface {nets[1]}")
            tun_vrf_s.append("no split-horizon")
            tun_vrf_s.append("no next-hop-self")
            tun_vrf_s.append("exit-af-interface ")

        tun_vrf_s.append("exit-address-family")
        tun_s[f"address-family ipv4 vrf {vrf} autonomous-system {nets[0]}"] = tun_vrf_s

    my_data["config"]["router eigrp DMVPN-EIGRP"] = tun_s

    return my_data

def fetch_site_data(config_file, site_number):
    try:
        data = {}
        hub = ""
        is_hub = False
        with open(config_file, 'r') as f:
            try:
                data = json.load(f)
                hub = data.get("hub", "")
            except json.JSONDecodeError:
                data = {}
                data["hub"] = f"site {site_number}"
                hub = f"site {site_number}"
                sites_len = 0
    except FileNotFoundError:
        data = {}
        data["hub"] = f"site {site_number}"
        hub = f"site {site_number}"
        sites_len = 0

    return data, hub == f"site {site_number}"

def create_global_config(router_id, intf_prefix):
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    ospf_s = []
    ospf_s.append(f"router-id {router_id}")
    ospf_s.append(f"exit")
    my_data["config"]["router ospf 1"] = ospf_s

    my_data["config"]["interface loopback0"] = []
    my_data["config"]["interface loopback0"].append(f"ip address {router_id} 255.255.255.255")
    my_data["config"]["interface loopback0"].append(f"ip ospf 1 area 0")
    my_data["config"]["interface loopback0"].append(f"exit")
    my_data["config"]["mpls ldp router-id Loopback0 force"] = []

    intf_s = []
    intf_s.append("ip address dhcp")
    intf_s.append("ip ospf 1 area 0")
    intf_s.append("mpls ip")
    intf_s.append("no shutdown")
    intf_s.append("exit")
    my_data["config"][f"interface {intf_prefix}1"] = intf_s

    my_data["network_info"]["loopback0"] = {
        "address": router_id,
        "mask": "255.255.255.255"
    }

    return my_data

def configure_site(sheet_file, config_file, sheet):
    sheet_data = read_sheet(sheet_file, sheet)

    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    ip_data = sheet_data["ip_data"]
    vrf_data = sheet_data["vrf_data"]
    tunnel_data = sheet_data["tunnel_data"]

    router_id = tunnel_data.iloc[0]['source']
    sn = sheet_data["md"].iloc[0]["site"]
    intf_prefix = sheet_data["md"].iloc[0]["intf_prefix"]

    data, is_hub = fetch_site_data(config_file, sn)
    my_data = create_global_config(router_id, intf_prefix)

    d_vrf = create_vrf(vrf_data, sn)
    my_data["config"].update(d_vrf["config"])
    my_data["network_info"].update(d_vrf["network_info"])

    d_ip = create_interface(ip_data, intf_prefix)
    my_data["config"].update(d_ip["config"])
    my_data["network_info"].update(d_ip["network_info"])

    d_bgp, data = create_mp_bgp_config(vrf_data, ip_data, data, router_id, sn)
    my_data["config"].update(d_bgp["config"])

    d_tunnel = create_tunnel_config(tunnel_data, data, is_hub)
    my_data["config"].update(d_tunnel["config"])
    my_data["network_info"].update(d_tunnel["network_info"])

    d_eigrp = create_tunnel_eigrp_config(vrf_data, tunnel_data, ip_data, is_hub)
    my_data["config"].update(d_eigrp["config"])
    my_data["network_info"].update(d_eigrp["network_info"])

    my_data["intf_prefix"] = intf_prefix

    data[f"site {sn}"] = my_data

    with open(config_file, 'w') as f:
        json.dump(data, f, indent=4)

    return data

def config_to_text(data, indent=0):
    lines = []
    prefix = "    " * indent

    if isinstance(data, dict):
        for key, value in data.items():
            lines.append(prefix + key)
            lines.extend(config_to_text(value, indent + 1))

    elif isinstance(data, list):
        for value in data:
            if isinstance(value, str):
                lines.append(prefix + value)
            else:
                lines.extend(config_to_text(value, indent))

    elif isinstance(data, str):
        lines.append(prefix + data)

    return lines

def create_or_update_config_files(data):
    del data["hub"]
    for site, site_data in data.items():
        config = site_data["config"]
        text = config_to_text(config)

        with open(f"site_text_configs/{site}.txt", "w", encoding="utf-8") as f:
            f.write("\n".join(text))


def main():
    if not os.path.exists("site_text_configs"):
        os.makedirs("site_text_configs")

    sheet_file = sys.argv[1]
    config_file = sys.argv[2]
    sheets = load_workbook(sheet_file).sheetnames

    data = {}
    for sheet in sheets:
        data.update(configure_site(sheet_file, config_file, sheet))

    create_or_update_config_files(data)

if __name__ == "__main__":
    main()