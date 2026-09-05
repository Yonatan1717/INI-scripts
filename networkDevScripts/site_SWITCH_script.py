import os
import sys
import json
from numpy import rint
import pandas as pd
from openpyxl import load_workbook
import ipaddress

SSH_DOMAIN = "lab.local"

def read_sheet(filename, sheet):
    df = pd.read_excel(
        filename,
        sheet_name=sheet,
        header=None
    )

    md_start = 17

    md_end = df.iloc[md_start:].isna().all(axis=1).idxmax()

    md = df.iloc[md_start:md_end, 0:5]
    md.columns = md.iloc[0]
    md = md[1:].reset_index(drop=True)

    swi_start = md_end + 1
    blank = df.iloc[swi_start:].isna().all(axis=1)
    if blank.any():
        swi_end = blank.idxmax()
    else:
        swi_end = len(df)

    swi_data = df.iloc[swi_start:swi_end, 0:9]
    swi_data.columns = swi_data.iloc[0]
    swi_data = swi_data[1:].reset_index(drop=True)

    # print(md)
    # print(swi_data)
    # exit()

    return {
        "md": md,
        "swi_data": swi_data,
    }


def enable_ssh(md, domain=SSH_DOMAIN):
    """
    Genererer lokal SSH-konfig
    brukernavn + passord.
    """
    my_data = {}
    my_data["config"] = {}
    my_data["network_info"] = {}

    if md.empty:
        return my_data

    row = md.iloc[0]

    username = row.get("brukernavn", "")
    password = row.get("passord", "")
    vty_lines = row.get("vty_lines", "0-4")

    if pd.isna(username) or pd.isna(password):
        return my_data

    username = str(username).strip()
    password = str(password).strip()

    if not username or not password:
        return my_data

    my_data["config"][f"ip domain name {domain}"] = []
    my_data["config"][
        f"username {username} privilege 15 secret {password}"
    ] = []
    my_data["config"]["crypto key generate rsa general-keys modulus 4096"] = []
    my_data["config"]["ip ssh version 2"] = []
    my_data["config"][f"line vty {' '.join(x.strip(' ') for x in vty_lines.split('-'))}"] = [
        "login local",
        "transport input ssh",
        "exit",
    ]

    return my_data


def global_config(md, swi_data):
    info = {}
    info["config"] = {}
    info["network_info"] = {}

    site = md.iloc[0]["site"]
    secret = md.iloc[0].get("secret", "")

    for idx, row in swi_data.iterrows():
        sw_id = row["SW"]
        mgmt_vlan = row["MGMT Vlan"]
        mgmt_ip = row["MGMT ip"]
        gatway = row["gateway"]
        mask = row["mask"]
        intf_prefix = row["intf_prefix"]

        if f"SW{sw_id}-SITE-{site}" not in info["config"]:
            info["config"][f"SW{sw_id}-SITE-{site}"] = {}

        info["config"][f"SW{sw_id}-SITE-{site}"][f"hostname SW{sw_id}-SITE-{site}"] = []
        info["config"][f"SW{sw_id}-SITE-{site}"][f"enable secret 9 {secret}"] = []

        ssh_config = enable_ssh(md)
        info["config"][f"SW{sw_id}-SITE-{site}"].update(ssh_config["config"])


        info["config"][f"SW{sw_id}-SITE-{site}"][f"vlan {mgmt_vlan}"] = [
            f"name MGMT_VLAN_{mgmt_vlan}",
            "exit"
        ]
        info["config"][f"SW{sw_id}-SITE-{site}"][f"interface vlan {mgmt_vlan}"] = [
            f"ip address {mgmt_ip} {mask}",
            "no shutdown",
            "exit"
        ]

        info["config"][f"SW{sw_id}-SITE-{site}"][f"interface {intf_prefix}0"] = [
            f"description Management interface for VLAN {mgmt_vlan}",
            "switchport mode access",
            f"switchport access vlan {mgmt_vlan}",
            "switchport port-security maximum 2",
            "switchport port-security violation restrict",
            "spanning-tree bpduguard enable",
            "spanning-tree portfast",
            "no shutdown",
            "exit"
        ]

        info["config"][f"SW{sw_id}-SITE-{site}"][f"ip default-gateway {gatway}"] = []
        info["config"][f"SW{sw_id}-SITE-{site}"][f"ntp server {gatway}"] = []


    return info


def config_vlan(swi_data, site):
    info = {} 
    info["config"] = {}
    info["network_info"] = {}

    for idx, row in swi_data.iterrows():
        sw_id = row["SW"]
        vlan_count = str(row["vlan-antall"])
        if "-" in vlan_count:
            vlan_info = [tuple(map(int, x.split("."))) for x in vlan_count.split("-")]
        else:
            vlan_info = [tuple(map(int, vlan_count.split(".")))]
            
        # print(vlan_info)
        # exit()

        intf_prefix = row["intf_prefix"]

        if f"SW{sw_id}-SITE-{site}" not in info["config"]:
            info["config"][f"SW{sw_id}-SITE-{site}"] = {}

        info["config"][f"SW{sw_id}-SITE-{site}"][f"vlan 999"] = [
            f"name NATIVE_UBRUKT",
            "exit"
        ]

        made = 0
        for vlan, antall in vlan_info:
            info["config"][f"SW{sw_id}-SITE-{site}"][f"vlan {vlan}"] = [
                f"name VLAN_{vlan}",
                "exit"
            ]

            rng = f"{made + 1}-{made + antall}" if antall > 1 else f"{made + 1}"
            range_or_not = "range " if antall > 1 else "" 

            info["config"][f"SW{sw_id}-SITE-{site}"][f"interface {range_or_not}{intf_prefix}{rng}"] = [
                f"description access port for VLAN {vlan}",
                "switchport mode access",
                f"switchport access vlan {vlan}",
                "switchport port-security maximum 2",
                "switchport port-security violation restrict",
                "spanning-tree bpduguard enable",
                "spanning-tree portfast",
                "no shutdown",
                "exit"
            ]

            made += antall


    return info


def config_trunk_and_dchp_snooping(swi_data, site):
    info = {} 
    info["config"] = {}
    info["network_info"] = {}

    for idx, row in swi_data.iterrows():
        sw_id = row["SW"]
        mgmg_vlan = int(row["MGMT Vlan"])
        vlan_count = str(row["vlan-antall"])
        if "-" in vlan_count:
            vlan_info = [tuple(map(int, x.split("."))) for x in vlan_count.split("-")]

            vlans = [vlan for vlan, antall in vlan_info]
            tot_antall_port = sum([int(antall) for vlan, antall in vlan_info])
        else:
            vlans = [int(vlan_count.split(".")[0])]
            antall = int(vlan_count.split(".")[1])
            tot_antall_port = antall

       
        if mgmg_vlan not in vlans:
            vlans.insert(0, mgmg_vlan)
            
        intf_prefix = row["intf_prefix"]
        num_ports = row["num_ports"]

        if f"SW{sw_id}-SITE-{site}" not in info["config"]:
            info["config"][f"SW{sw_id}-SITE-{site}"] = {}

        to_lan = num_ports - 1
        to_core = num_ports - 2

        info["config"][f"SW{sw_id}-SITE-{site}"][f"ip dhcp snooping"] = []
        info["config"][f"SW{sw_id}-SITE-{site}"][f"ip dhcp snooping vlan {','.join(map(str, vlans))}"] = []
        info["config"][f"SW{sw_id}-SITE-{site}"][f"no ip dhcp snooping information option"] = []
        info["config"][f"SW{sw_id}-SITE-{site}"][f"ip arp inspection vlan {','.join(map(str, vlans))}"] = []

        info["config"][f"SW{sw_id}-SITE-{site}"][f"interface {intf_prefix}{to_core}"] = [
            f"description uplink trunk port for VLAN {','.join(map(str, vlans))}",
            "switchport trunk encapsulation dot1q",
            "switchport trunk native vlan 999",
            "switchport mode trunk",
            f"switchport trunk allowed vlan {','.join(map(str, vlans))}",
            f"ip dhcp snooping trust",
            f"ip arp inspection trust",
            "no shutdown",
            "exit"
        ]


        info["config"][f"SW{sw_id}-SITE-{site}"][f"interface {intf_prefix}{to_lan}"] = [
            f"description downlink trunk port for VLAN {','.join(map(str, vlans))} om ikke brukt skal det brukes shutdown på porten",
            "switchport trunk encapsulation dot1q",
            "switchport trunk native vlan 999",
            "switchport mode trunk",
            f"switchport trunk allowed vlan {','.join(map(str, vlans))}",
            "no shutdown",
            "exit"
        ]
    
        ports_left = num_ports - tot_antall_port - 3
        if ports_left > 0:
            start_int = tot_antall_port + 1
            end_int = num_ports - 3
            range_or_not = "range " if start_int != end_int else ""

            rng = f"{start_int}-{end_int}" if start_int != end_int else f"{start_int}"

            info["config"][f"SW{sw_id}-SITE-{site}"][f"interface {range_or_not}{intf_prefix}{rng}"] = [
                f"description ubrukt port access ports",
                "shutdown",
                "exit"
            ]


    return info


def update_site_config(data,swi_data, sn, conf):
    for sw_id in swi_data["SW"].unique():
        if f"SW{sw_id}-SITE-{sn}" not in data[f"site {sn}"]["config"]:
            data[f"site {sn}"]["config"][f"SW{sw_id}-SITE-{sn}"] = {}

        data[f"site {sn}"]["config"][f"SW{sw_id}-SITE-{sn}"].update(conf["config"][f"SW{sw_id}-SITE-{sn}"])

    return data


def fetch_site_data(config_file):
    try:
        data = {}

        with open(config_file, "r") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                data = {}

    except FileNotFoundError:
        data = {}

    return data


def create_site_sw_config(file, sheet, config_file):
    sheet_data = read_sheet(file, sheet)
    data = fetch_site_data(config_file)

    md = sheet_data["md"]
    swi_data = sheet_data["swi_data"]

    sn = md.iloc[0]["site"]
    data[f"site {sn}"] = {}
    data[f"site {sn}"]["config"] = {}

    global_conf = global_config(md, swi_data)
    data = update_site_config(data, swi_data, sn, global_conf)

    vlan_conf = config_vlan(swi_data, sn)
    data = update_site_config(data, swi_data, sn, vlan_conf)

    trunk_and_dchp_snooping_conf = config_trunk_and_dchp_snooping(swi_data, sn)
    data = update_site_config(data, swi_data, sn, trunk_and_dchp_snooping_conf)

    with open(config_file, "w") as f:
        f.write(json.dumps(data, indent=4))


    return data


def config_to_text(data, indent=0):
    lines = []
    prefix = "    " * indent

    if isinstance(data, dict):
        for key, value in data.items():
            lines.append("!")
            lines.append(prefix + key)
            lines.extend(config_to_text(value, indent + 1))

    elif isinstance(data, list):
        for value in data:
            if isinstance(value, str):
                lines.append(prefix + value)
            else:
                lines.extend(config_to_text(value, indent))

    elif isinstance(data, str):
        lines.append(prefix + data)
        
    lines.append("!")
  
    return lines


def create_or_update_config_files(data):
    if not os.path.exists("siteSwichTextConfigs"):
        os.makedirs("siteSwichTextConfigs")

        
    for site, site_data in data.items():
        if not os.path.exists(f"siteSwichTextConfigs/site_{site}"):
            os.makedirs(f"siteSwichTextConfigs/site_{site}")

        configs = site_data["config"]

        for sw_name, config in configs.items():
            text = config_to_text(config)

            with open(
                f"siteSwichTextConfigs/site_{site}/{sw_name}.txt",
                "w",
                encoding="utf-8"
            ) as f:
                f.write("\n".join(text))
            

    print(f"Text versjon av config for svitjer i site {site}, har blitt lagret i siteSwichTextConfigs/site_{site}/{sw_name}.txt")


def create_sw_configs_main(file, config_file="site_switch_config.json"):   

    sites_sheets = load_workbook(file).sheetnames
    
    for sheet in sites_sheets:
        data = create_site_sw_config(file, sheet, config_file)
    
    create_or_update_config_files(data)
   
        
def main():
    file = sys.argv[1]
    config_file = "site_switch_config.json" if len(sys.argv) < 3 else sys.argv[2]
    create_sw_configs_main(file, config_file)


if __name__ == "__main__":
    main()